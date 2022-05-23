import openpyxl

標準功效 = [
    ["補脾肺腎", "補脾氣", "補肺氣", "補肺陰", "補腎陰", "補脾陰"],
    ["補脾和胃", "補脾氣", "降胃氣"],
    ["補脾止瀉", "止瀉", "補脾陽", "袪脾寒"],
    ["補脾益肺", "補脾氣", "補肺氣"],
    ["補脾益氣", "補脾氣"],
    ["補氣健脾", "補脾氣"],
    ["補肺氣", "補肺氣"],
    ["補氣升陽", "補脾氣", "補肺氣", "補中氣"],
    ["補血", "補心血", "補肝血", "補脾血"],
    ["補中益氣", "補脾氣"],
    ["補腎助陽", "補腎陽"],

    ["排膿", "清熱解毒", "排膿"],
    ["排膿通乳", "消散癰腫"],
    ["破氣除痞", "清胃熱", "清大腸熱", "消心下痞"],
    ["破血消癥", "活血", "消癥"],
    ["平喘", "平喘"],
    ["平抑肝陽", "平肝陽", "抑肝陽"],

    ["明目", "清肝熱明目"],
    ["明目退翳", "清肝熱明目"],

    ["發表透疹", "發散表熱", "透疹"],
    ["發表散風", "發散表風"],
    ["發汗透疹", "發散表風", "透疹"],
    ["發汗解表", "袪風寒", "解表發汗", "發散表寒"],

    ["大補元氣", "補元氣"],
    ["定喘嗽", "平喘"],

    ["托瘡生肌", "托瘡生肌"],
    ["通絡", "通經絡"],
    ["通絡止痛", "通經絡", "止痛"],
    ["通利小便", "利尿通淋"],
    ["通經活絡", "通經絡"],
    ["通竅", "通鼻竅"],
    ["通血脈", "散淤"],
    ["通乳", "通乳"],
    ["通陽散寒", "溫經通陽"],
    ["退虛熱", "清心虛熱", "清肝虛熱", "清肝疳熱"],
    ["透疹消瘡", "透疹", "消瘡"],
    ["透疹止癢", "透疹", "止癢"],
    ["調經止痛", "止痛"],

    ["利尿", "利尿通淋"],
    ["利尿通淋", "利尿通淋"],
    ["利氣散結", "行肺氣", "消散癰腫"],
    ["利小便", "利尿通淋"],
    ["利水退腫", "消浮腫"],
    ["利水通淋", "利心濕", "利小腸濕", "利膀胱濕", "利尿通淋", "利胃濕"],
    ["利水清熱", "利肺濕", "清肺熱", "清胃熱"],
    ["利水消腫", "利肺濕", "消水腫", "利脾濕", "利膀胱濕", "利腎濕"],
    ["利水滲濕", "利脾濕"],
    ["利濕", "利肺濕", "利大腸濕"],
    ["利濕健脾", "利脾濕"],
    ["利濕除痹", "利脾濕", "除痺"],
    ["利咽", "通利咽喉"],
    ["利咽開音", "清肺熱", "通利咽喉"],
    ["理氣調中", "行脾氣"],
    ["理氣寬中", "行脾氣"],
    ["理氣安胎", "行脾氣", "行胃氣", "安胎"],
    ["斂肺定喘", "平喘"],
    ["斂肺止咳", "止咳"],
    ["斂汗", "止汗"],
    ["涼血退蒸", "清肝熱涼血", "清胃熱涼血", "清大腸熱涼血"],
    ["涼血利咽", "清心熱涼血", "清胃熱涼血", "通利咽喉"],
    ["涼血活血透疹", "清心熱涼血", "清肝熱涼血", "活心血", "活肝血", "透疹"],
    ["涼血消斑", "清肝熱涼血", "清肺熱涼血", "清胃熱涼血", "消斑"],
    ["涼血止血", "清肝熱涼血", "清胃熱涼血", "止血"],

    ["固精縮尿", "縮尿", "補腎精"],
    ["攻毒散結", "解肝毒", "消散癰腫"],

    ["開提肺氣", "開提肺氣"],
    ["寬胸散結", "消心下痞"],

    ["緩和藥性", "調和藥性"],
    ["緩急止痛", "調和藥性", "止痛"],
    ["化痰消積", "化痰"],
    ["化痰散結", "化痰", "消散癰腫"],
    ["化濕", "燥脾濕", "燥胃濕"],
    ["化濕開胃", "燥脾濕", "燥胃濕", "燥腎濕", "補脾氣"],
    ["化濕和中", "燥肺濕"],
    ["化濕解表", "燥脾濕", "解表發汗"],
    ["化濕祛暑", "燥脾濕", "袪暑"],
    ["化濕行氣", "行脾氣", "燥脾濕", "燥胃濕", "行胃氣"],
    ["活血消癥", "活肝血", "活胃血", "消癥"],
    ["活血消腫", "活肝血", "活胃血"],
    ["活血行氣止痛", "活心血", "活肝血", "活脾血", "行肝氣", "行脾氣"],
    ["活血止痛", "活血", "止痛"],
    ["回陽", "回陽"],
    ["和中健脾", "補脾氣", "降胃氣"],

    ["健脾", "補脾氣"],
    ["健脾化濕", "補脾氣", "利脾濕"],
    ["降逆止嘔", "止嘔", "降胃氣", "降肺氣", "清胃熱"],
    ["降火生津", "生津"],
    ["降氣化痰", "降肺氣", "化痰"],
    ["解表", "發散表寒"],
    ["解表散風", "發散表風"],
    ["解毒", "解毒", "解肺毒"],
    ["解毒療瘡", "解肝毒", "消瘡", "解心毒", "療瘡"],
    ["解毒散結", "解蛇毒", "消散癰腫"],
    ["解暑化濕", "燥脾濕"],
    ["截瘧", "截瘧"],

    ["清肺化痰", "清肺熱", "化痰"],
    ["清肺化痰止咳", "清肺熱", "化痰", "止咳"],
    ["清利頭目", "袪目痛", "袪頭痛"],
    ["清疳熱", "清胃疳熱"],
    ["清肝瀉火", "瀉肝火"],
    ["清解暑熱", "清膀胱熱"],
    ["清虛熱", "清肝虛熱", "清膽虛熱", "清腎虛熱"],
    ["清心潤肺", "清心熱", "補肺陰"],
    ["清暑化濕", "清心熱", "燥脾濕"],
    ["清熱", "清心熱"],
    ["清熱排膿", "清熱解毒", "排膿"],
    ["清熱涼血", "清肝熱", "清肝熱涼血", "清肝陰涼血", "清胃熱涼血",
        "清胃陰涼血", "清心陰涼血", "清肺陰涼血", "清心熱", "清肺熱"],
    ["清熱化痰", "清肺熱", "化痰", "清肝熱"],
    ["清熱解毒", "清熱解毒", "清心熱", "解心毒", "清肝熱", "解肝毒",
        "清胃熱", "解胃毒", "清肺熱", "解肺毒", "清大腸熱", "解大腸毒"],
    ["清熱解暑", "清心熱", "清膀胱熱"],
    ["清熱息風", "清肺熱", "清膀胱熱"],
    ["清熱除煩", "清心熱", "除煩"],
    ["清熱生津", "清肺熱", "清胃熱", "生津"],
    ["清熱燥濕", "清大腸熱", "燥大腸濕", "清膀胱熱", "燥膀胱濕", "清肝熱",
        "燥肝濕", "清胃熱", "燥胃濕", "清心熱", "清肺熱", "燥肺濕"],
    ["清熱養陰", "清肺熱", "補肺陰", "補腎陰"],
    ["祛風痰", "息肝風"],
    ["祛風濕", "祛風濕"],
    ["袪風通絡", "袪風", "通經絡"],
    ["祛風解痙", "止痙", "息肝風"],
    ["祛風止痛", "祛風", "止痛"],
    ["祛風除濕", "袪風濕"],
    ["祛風散寒", "發散表寒"],
    ["袪風,散寒止痛", "袪風寒", "止痛"],
    ["袪痰降氣止咳", "降肺氣", "止咳", "化痰"],
    ["袪痰止咳", "化痰"],
    ["袪暑", "袪暑"],
    ["袪暑", "袪暑"],
    ["祛濕斂瘡", "收斂生肌"],
    ["驅蟲", "驅蟲"],
    ["驅蟲消積", "消食積"],
    ["強筋骨", "補腎精"],

    ["消痞散結", "消痞散結"],
    ["消痰", "化痰"],
    ["消積", "消食積"],
    ["消積通便", "消食積", "潤腸燥"],
    ["消滯", "消食積"],
    ["消腫排膿", "消散癰腫"],
    ["消腫潰堅", "消浮腫"],
    ["消腫生肌", "去腐生肌"],
    ["消腫散結", "消腫散結"],
    ["消食除脹", "消食積"],
    ["消癰排膿", "消散癰腫", "排皮膚膿"],
    ["消癰散結", "消散癰腫"],
    ["宣肺袪痰利咽", "宣肺氣", "通利咽喉", "化痰"],
    ["宣散風熱", "清肺熱"],
    ["洩肺氣", "洩肺氣"],
    ["瀉肺平喘", "清肺熱", "平喘"],
    ["瀉肝膽火", "瀉膽火", "瀉肝火"],
    ["瀉火解毒", "瀉肝火", "瀉胃火", "瀉大腸火", "解肝毒", "解胃毒",
        "解大腸毒", "瀉心火", "瀉肺火", "解心毒", "解肺毒"],
    ["瀉下", "潤腸燥"],
    ["瀉下逐水", "瀉大腸的積滯", "逐水"],
    ["瀉水逐飲", "逐水"],
    ["行氣", "行脾氣", "行肺氣", "行胃氣"],
    ["行氣調中止痛", "行脾氣", "止腹痛"],
    ["行氣利水", "利大腸氣", "利膀胱濕", "行胃氣"],
    ["行氣寬中", "行脾氣", "止嘔", "行胃氣", "寬中"],
    ["息風定驚", "息肝風"],
    ["息風止痙", "止痙", "息肝風"],
    ["泄心火", "瀉心火"],
    ["下氣", "降肺氣"],

    ["止痹痛", "止痹痛"],
    ["止帶", "止帶"],
    ["止痛", "止痛"],
    ["止咳平喘", "止咳", "平喘"],
    ["止汗", "補衛氣", "固表止汗"],
    ["止痙", "止痙"],
    ["止血", "止血"],
    ["止嘔", "降胃氣", "止嘔"],
    ["治癬", "治癬"],
    ["助腎陽", "補腎陽"],
    ["逐痰", "逐痰"],

    ["除煩止渴", "除煩", "生津"],
    ["除煩止嘔", "除煩", "止嘔"],
    ["除疳熱", "清肝疳熱", "清胃疳熱"],
    ["除骨蒸", "除骨蒸"],
    ["除濕熱", "清胃熱", "清大腸熱", "燥胃濕", "燥大腸濕"],
    ["炒炭止血", "止血"],

    ["生津止渴", "生津"],
    ["升發清陽", "清肝熱", "升肝陽"],
    ["升舉陽氣", "升脾陽"],
    ["殺蟲利尿", "殺蟲", "利尿通淋"],
    ["殺蟲攻毒", "殺蟲", "解蛇毒"],
    ["疏肝理氣", "行肝氣"],
    ["疏散風熱", "清肺熱", "發散表熱", "發散表風"],
    ["勝濕止痛", "止痛", "袪風濕", "止痹痛"],

    ["潤肺止咳", "補肺陰", "止咳"],
    ["潤肺止咳化痰", "補肺陰", "止咳", "化痰"],
    ["潤腸", "潤腸燥"],
    ["潤腸通便", "潤腸燥", "潤大腸燥"],
    ["柔肝止痛", "止痛"],
    ["軟堅消痰", "軟堅散結"],
    ["軟堅散結", "軟堅散結"],

    ["燥濕", "燥肺濕"],
    ["燥濕痰", "化痰"],
    ["燥濕利水", "燥脾濕"],
    ["燥濕化痰", "燥脾濕", "化痰"],
    ["燥濕健脾", "燥脾濕", "補脾氣"],
    ["燥濕止帶", "燥胃濕", "止帶"],
    ["燥濕溫中", "燥脾濕", "補脾陽"],
    ["滋陰潛陽", "補肝陰", "補腎陰"],

    ["縮尿", "縮尿"],
    ["散寒止痛", "發散表寒", "止痛", "袪肝寒"],
    ["澀腸止瀉", "澀腸"],

    ["益肝腎", "補肝血"],
    ["益精血", "補腎精"],
    ["益氣復脈", "補心氣"],
    ["益氣養陰", "補脾氣", "補肺氣", "補肺陰", "補腎陰", "補脾陰"],
    ["益衛固表", "補衛氣", "固表止汗"],
    ["養血斂陰", "補肝血", "斂肝陰"],
    ["養血安神", "補脾血", "安神"],
    ["養心安神", "安神", "補心血"],
    ["養心益脾", "補心氣"],
    ["養陰生津", "補心陰", "補肝陰", "補肺陰", "生津"],
    ["養陰潤燥", "補肺陰", "補胃陰"],

    ["溫脾胃", "補脾陽"],
    ["溫脾止瀉", "止瀉", "補脾陽", "袪脾寒"],
    ["溫肺化痰", "袪肺寒", "化痰"],
    ["溫肺化飲", "化痰", "補肺陽"],
    ["溫肺止咳", "補肺陽", "止咳"],
    ["溫精通陽", "補經絡陽"],
    ["溫經止血", "補經絡陽", "止血"],
    ["溫中", "補脾陽"],
    ["溫中止瀉", "止瀉", "補脾陽"],
    ["溫中止嘔", "止嘔", "補脾陽"],
    ["溫胃", "袪脾寒"],
    ["外用消腫止痛", "消散癰腫", "止痛"],

    ["湧吐", "湧吐"],

    ["安胎", "安胎"],
    ["安神", "安神"],
    ["安神增智", "補心氣", "安神"]
]

藥材功效 = [
    ["麻黃", "透疹", "發散表寒", "通鼻竅", "利三焦濕", "利皮毛濕", "利肺濕", "消水腫", "平喘", "止癢", "宣肺氣"],
    ["桂枝", "補心陽", "發散表寒", "祛風寒", "溫經通陽", "補經絡陽", "解表發汗"],
    ["杏仁", "降肺氣", "止咳", "潤腸燥", "平喘"],
    ["炙甘草", "補心氣", "補脾氣", "降胃氣"],
    ["白芍", "緩急", "補肝血", "斂肝陰", "抑肝陽", "平肝陽", "止痛"],
    ["生薑", "祛風寒", "補肺陽", "補心陽", "發散表寒", "解胃毒"],
    ["大棗", "補胃氣", "調和藥性", "補脾血", "養心安神", "補脾氣"],
    ["羌活", "袪膀胱經風寒", "燥膀胱濕", "止痛"],
    ["防風", "祛風寒", "發散表寒", "止痙", "止痛", "止瀉"],
    ["蒼朮", "祛風濕", "燥脾濕", "補脾氣"],
    ["細辛", "發散裡寒", "通膀胱竅", "通腎竅", "化痰", "補肺陽"],
    ["川芎", "行肝氣", "發散表寒", "止痛", "活血", "祛風寒"],
    ["白芷", "消散癰腫", "發散表風", "止帶", "通鼻竅", "燥胃濕", "止痛"],
    ["黃芩", "解肝毒", "瀉肝火", "清三焦熱涼血", "清肝熱涼血", "瀉心火", "燥胃濕", "清膽熱", "清膽熱涼血", "清三焦熱", "燥三焦濕", "燥肺濕",
        "清心熱涼血", "止血", "清肺熱涼血", "清胃熱涼血", "清胃熱", "燥膽濕", "瀉大腸火", "燥大腸濕", "清肺熱", "瀉胃火", "清大腸熱", "安胎", "清腎熱"],
    ["生地黃", "清肝熱", "補肝陰涼血", "補腎陰", "補心陰涼血", "補腎陰涼血",
        "補肝陰", "補肺陰", "生津", "清腎熱", "補肺陰涼血", "清心熱"],
    ["甘草", "調和藥性", "清熱解毒", "化痰", "解胃毒", "清胃熱", "補脾氣", "止咳", "補肺陰"],
    ["獨活", "發散表寒", "止痹痛", "祛風濕"],
    ["藁本", "祛風濕", "發散表寒", "止痹痛", "止痛"],
    ["蔓荊子", "祛頭痛", "發散表風", "祛目痛", "發散表熱"],
    ["紫蘇", "發散表寒", "止嘔", "行脾氣"],
    ["香附", "行肝氣", "止經痛", "調月經"],
    ["陳皮", "行脾氣", "燥脾濕", "降胃氣", "化痰", "止嘔"],
    ["乾薑", "補脾陽", "補經絡陽", "溫月經", "回陽", "補肺陽", "止血", "化痰"],
    ["半夏", "降胃氣", "止嘔", "消心下痞", "散心下結", "化痰", "消痞散結", "燥脾濕"],
    ["五味子", "止血", "補胃氣", "歛肺氣", "生津", "收斂固澀", "止汗", "補腎陰", "開心竅"],
    ["紫菀", "化痰", "止咳"],
    ["百部", "殺蟲", "止咳", "潤肺燥"],
    ["白前", "降肺氣", "化痰", "止咳"],
    ["桔梗", "宣肺氣", "開提肺氣", "清熱解毒", "排皮膚膿", "祛咽喉痛",
        "排肺膿", "通利咽喉", "排膿", "手太陰肺經", "化痰"],
    ["橘紅", "化痰", "行脾氣", "燥脾濕"],
    ["荊芥", "發散表風", "消瘡", "止血", "透疹"],
    ["金銀花", "清肺熱", "消散癰腫", "解心毒", "清胃熱", "清胃虛熱", "祛風熱", "解肺毒", "清心熱"],
    ["連翹", "發散表熱", "清心熱", "清肺熱", "消散癰腫", "解心毒"],
    ["牛蒡子", "降肺氣", "化痰", "止咳"],
    ["薄荷", "祛頭痛", "宣肺熱", "祛咽喉痛", "透疹", "祛風熱", "行肝氣"],
    ["淡豆豉", "除煩", "發散表熱"],
    ["淡竹葉", "利尿通淋", "清心熱", "除煩", "利心濕", "利小腸濕"],
    ["蘆根", "生津", "清肺熱", "除煩", "止嘔", "清胃熱"],
    ["菊花", "平肝熱", "解肝毒", "明目", "清肝虛熱", "祛風熱"],
    ["石膏", "清胃熱", "收斂生肌", "生津", "瀉胃火", "除煩", "清肺熱"],
    ["葛根", "清胃虛熱", "燥肝濕", "生津", "升脾陽", "透疹", "止瀉", "透發麻疹"],
    ["柴胡", "行肝氣", "清膽熱", "升肝陽", "清肝熱", "足厥陰肝經", "手厥陰心包絡經", "手少陽三焦經", "足少陽膽經"],
    ["升麻", "透疹", "解胃毒", "升胃陽", "升脾陽", "清胃熱", "發散表熱", "足太陰脾經", "手太陰肺經"],
    ["枳殼", "清胃熱", "清大腸熱", "消食積", "消心下痞", "化痰"],
    ["前胡", "降肺氣", "清肺熱", "化痰"],
    ["茯苓", "利脾濕", "補脾氣", "養心安神"],
    ["人參", "補脾氣", "補元氣", "養心安神", "補肺氣", "補心氣", "生津"],
    ["木香", "行脾氣", "止腹痛"],
    ["附子", "回陽", "止痛", "祛風寒", "燥腎濕", "發散裡寒", "補腎陽"],
    ["萎蕤", "補脾陰", "補胃陰", "生津"],
    ["蔥", "發散表寒", "溫經通陽", "安回", "解肺毒"],
    ["白薇", "清胃熱涼血", "清肝熱涼血", "利尿通淋", "補肝陰涼血", "消瘡", "清肝熱", "解肝毒", "補胃陰涼血"]
]


# open excel
data = openpyxl.load_workbook("原始藥效.xlsx")


def systemeffect(row):
    ori_effect = sheet.cell(row, 4).value
    std_col = 5
    bool_system = False
    for check in range(len(標準功效)):
        if ori_effect == 標準功效[check][0]:
            bool_system = True
            for i in range(1, len(標準功效[check])):
                sheet.cell(row, std_col).value = 標準功效[check][i]
                std_col += 1
            break
    return bool_system


def compareeffect(row, index):
    # output is stored at col 5
    std_col = 5

    # original effect's char
    for char_ori in sheet.cell(row, 4).value:
        # std effect
        for str_std in 藥材功效[index]:
            for char_std in str_std:
                # if there is a char is the same between original effect & std effect
                if char_ori == char_std:
                    # if the std effect is repeat at the front cell, continue
                    if(std_col > 5 and str_std == sheet.cell(row, std_col-1).value):
                        continue
                    # else, std effect is stored at col (>=5)
                    sheet.cell(row, std_col).value = str_std
                    std_col += 1


for sheet in data.worksheets:
    # loop the whole file
    for row in range(2, sheet.max_row, 1):
        bool_system = systemeffect(row)

        # medicine's name
        if(bool_system == False):
            checkName = sheet.cell(row, 3).value
            for check in range(len(藥材功效)):
                # 藥材功效[i][0] is the medicine name
                if checkName == 藥材功效[check][0]:
                    compareeffect(row, check)

data.save('原始藥效.xlsx')